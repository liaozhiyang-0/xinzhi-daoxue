from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal


class ResearchTabularReadError(ValueError):
    """Raised when a declared local tabular format cannot be read safely."""


TabularFormat = Literal["csv", "tsv", "json", "xlsx", "parquet"]


def read_tabular_rows(
    path: Path, format_name: TabularFormat
) -> tuple[list[str], list[dict[str, str]]]:
    """Read a bounded, row-oriented table without inferring a different format.

    CSV/TSV/JSON use the standard library. XLSX and Parquet are optional adapters;
    when their dependency is absent, the error is explicit so the quality gate can
    report a missing runtime capability instead of silently treating the file as
    another format.
    """

    if format_name in {"csv", "tsv"}:
        return _read_delimited(path, "\t" if format_name == "tsv" else ",")
    if format_name == "json":
        return _read_json(path)
    if format_name == "xlsx":
        return _read_xlsx(path)
    if format_name == "parquet":
        return _read_parquet(path)
    raise ResearchTabularReadError(f"unsupported_local_format:{format_name}")


def _read_delimited(
    path: Path, delimiter: str
) -> tuple[list[str], list[dict[str, str]]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            if not reader.fieldnames:
                raise ResearchTabularReadError("dataset_header_missing")
            columns = [str(item) for item in reader.fieldnames]
            if len(columns) != len(set(columns)):
                raise ResearchTabularReadError("dataset_header_has_duplicates")
            rows = [
                {column: _stringify_value(value) for column, value in row.items()}
                for row in reader
            ]
            return columns, rows
    except UnicodeDecodeError as exc:
        raise ResearchTabularReadError("dataset_text_encoding_unsupported") from exc


def _read_json(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ResearchTabularReadError("json_dataset_invalid") from exc
    raw_rows = payload.get("rows") if isinstance(payload, dict) else payload
    if not isinstance(raw_rows, list) or not all(
        isinstance(row, dict) for row in raw_rows
    ):
        raise ResearchTabularReadError("json_dataset_must_be_a_row_array")
    columns = sorted({str(key) for row in raw_rows for key in row})
    if len(columns) != len(set(columns)):
        raise ResearchTabularReadError("dataset_header_has_duplicates")
    rows = [
        {column: _stringify_value(row.get(column, "")) for column in columns}
        for row in raw_rows
    ]
    return columns, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    workbook: Any = None
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ResearchTabularReadError(
            "optional_dependency_missing:openpyxl_for_xlsx"
        ) from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        header = next(values, None)
        if header is None:
            raise ResearchTabularReadError("dataset_header_missing")
        columns = [_header_value(value) for value in header]
        if not all(columns):
            raise ResearchTabularReadError("dataset_header_contains_empty_name")
        if len(columns) != len(set(columns)):
            raise ResearchTabularReadError("dataset_header_has_duplicates")
        rows: list[dict[str, str]] = []
        for values_row in values:
            row = {
                column: _stringify_value(
                    values_row[index] if index < len(values_row) else ""
                )
                for index, column in enumerate(columns)
            }
            if any(value for value in row.values()):
                rows.append(row)
        return columns, rows
    except ResearchTabularReadError:
        raise
    except (OSError, ValueError, TypeError) as exc:
        raise ResearchTabularReadError("xlsx_dataset_invalid") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _read_parquet(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from pyarrow import parquet  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ResearchTabularReadError(
            "optional_dependency_missing:pyarrow_for_parquet"
        ) from exc
    try:
        table = parquet.read_table(path)
        columns = [str(item) for item in table.column_names]
        if len(columns) != len(set(columns)):
            raise ResearchTabularReadError("dataset_header_has_duplicates")
        rows = [
            {column: _stringify_value(row.get(column, "")) for column in columns}
            for row in table.to_pylist()
        ]
        return columns, rows
    except ResearchTabularReadError:
        raise
    except (OSError, ValueError, TypeError) as exc:
        raise ResearchTabularReadError("parquet_dataset_invalid") from exc


def _header_value(value: object) -> str:
    return _stringify_value(value).strip()


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).strip()
